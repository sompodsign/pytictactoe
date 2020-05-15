import sys
import openpyxl

name = sys.argv[1]
blank_start = sys.argv[2]
blank_length = sys.argv[3]

wb = openpyxl.load_workbook(name)
sheet = wb.active

# Read the spreadsheet and get list of row data lists
print('Reading spreadsheet data...')
rows = []
max_row = sheet.max_row
max_column = sheet.max_column
for row in range(1, sheet.max_row+1):
    data = []
    for cell in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=cell).value
        data.append(cell_value)
    rows.append(data)

wb = openpyxl.Workbook()
sheet = wb.active

print('Inserting blank rows...')
for row in range(1, blank_start):
    for cell in range(1, max_column + 1):
        sheet.cell(row=row, column=cell).value = rows[row -1][cell - 1]