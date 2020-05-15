# updateProduce.py - corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Updated Prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

#Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    product = sheet.cell(row= rowNum, column = 1).value
    if product in PRICE_UPDATES:
        sheet.cell(row = rowNum, column = 2).value = PRICE_UPDATES[product]
wb.save('updatedProduceSales.xlsx')